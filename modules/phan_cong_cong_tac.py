import os
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from flask import (
    Blueprint, render_template, request, jsonify, redirect, 
    url_for, flash, send_file, current_app
)
from flask_login import login_required, current_user
from sqlalchemy import or_, and_, func

from extensions import db
from models import PhanCongCongTac, User, PhongBan, ThongTinNguoiLaoDong, LinhVuc, LichSuPhuTrach

phan_cong_bp = Blueprint('phan_cong', __name__, url_prefix='/phan-cong-cong-tac')

# ----------------------------------------------------------------------
# 1. TRANG DANH SÁCH & LỊCH PHÂN CÔNG CÔNG TÁC
# ----------------------------------------------------------------------
@phan_cong_bp.route('/', methods=['GET'])
@login_required
def index():
    page = request.args.get('page', 1, type=int)
    keyword = request.args.get('keyword', '').strip()
    trang_thai = request.args.get('trang_thai', '').strip()
    tu_ngay = request.args.get('tu_ngay', '').strip()
    den_ngay = request.args.get('den_ngay', '').strip()

    query = PhanCongCongTac.query

    # Nếu không phải Admin/Quản lý thì chỉ xem lịch của chính mình hoặc do mình tạo
    is_admin_or_mgr = (
        getattr(current_user, 'is_admin', False) or 
        getattr(current_user, 'can_approve', False) or 
        getattr(current_user, 'is_controller', False)
    )
    if not is_admin_or_mgr:
        query = query.filter(
            or_(
                PhanCongCongTac.ma_nhan_vien == current_user.ma_nhan_vien,
                PhanCongCongTac.nguoi_tao == current_user.ma_nhan_vien
            )
        )

    if keyword:
        query = query.filter(
            or_(
                PhanCongCongTac.ho_ten.ilike(f'%{keyword}%'),
                PhanCongCongTac.ma_nhan_vien.ilike(f'%{keyword}%'),
                PhanCongCongTac.noi_cong_tac.ilike(f'%{keyword}%'),
                PhanCongCongTac.noi_dung_cong_tac.ilike(f'%{keyword}%')
            )
        )

    if trang_thai:
        query = query.filter(PhanCongCongTac.trang_thai == trang_thai)

    if tu_ngay:
        try:
            d_tu = datetime.strptime(tu_ngay, '%Y-%m-%d').date()
            query = query.filter(PhanCongCongTac.ngay_bat_dau >= d_tu)
        except ValueError:
            pass

    if den_ngay:
        try:
            d_den = datetime.strptime(den_ngay, '%Y-%m-%d').date()
            query = query.filter(PhanCongCongTac.ngay_ket_thuc <= d_den)
        except ValueError:
            pass

    pagination = query.order_by(PhanCongCongTac.created_at.desc()).paginate(
        page=page, per_page=15, error_out=False
    )
    ds_phan_cong = pagination.items

    # Loại bỏ ds_can_bo = ThongTinNguoiLaoDong.query.all() để tránh tốn RAM/Query time (vì đã có AJAX Select2)
    ds_phong_ban = PhongBan.query.all() if hasattr(PhongBan, 'query') else []

    # Lấy danh sách tất cả các lĩnh vực công tác từ DB
    ds_linh_vuc = LinhVuc.query.order_by(LinhVuc.ten_linh_vuc.asc()).all() if hasattr(LinhVuc, 'query') else []

    return render_template(
        'phan_cong_cong_tac/index.html',
        ds_phan_cong=ds_phan_cong,
        pagination=pagination,
        ##ds_can_bo=ds_can_bo,
        ds_phong_ban=ds_phong_ban,
        ds_linh_vuc=ds_linh_vuc,
        keyword=keyword,
        trang_thai=trang_thai,
        tu_ngay=tu_ngay,
        den_ngay=den_ngay
    )

# ----------------------------------------------------------------------
# API LẤY LĨNH VỰC ĐANG PHỤ TRÁCH CỦA CÁN BỘ (Dựa trên lich_su_phu_trach)
# ----------------------------------------------------------------------
@phan_cong_bp.route('/get-linh-vuc-can-bo/<ma_nv>', methods=['GET'])
@login_required
def get_linh_vuc_can_bo(ma_nv):
    try:
        records = db.session.query(LinhVuc.id, LinhVuc.ten_linh_vuc)\
            .join(LichSuPhuTrach, LichSuPhuTrach.linh_vuc_id == LinhVuc.id)\
            .filter(LichSuPhuTrach.ma_nhan_vien == ma_nv, LichSuPhuTrach.ngay_ket_thuc.is_(None))\
            .all()
        
        linh_vuc_data = [{"id": r[1], "text": r[1]} for r in records]
        return jsonify({'success': True, 'linh_vuc': linh_vuc_data})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@phan_cong_bp.route('/api/search-can-bo', methods=['GET'])
@login_required
def search_can_bo():
    query_str = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 20
    offset = (page - 1) * per_page

    base_query = ThongTinNguoiLaoDong.query

    if query_str:
        # Tạo danh sách các điều kiện tìm kiếm trên các cột Chuỗi hợp lệ
        conditions = [
            ThongTinNguoiLaoDong.ho_ten.ilike(f"%{query_str}%"),
            ThongTinNguoiLaoDong.ma_nhan_vien.ilike(f"%{query_str}%")
        ]

        # Kiểm tra nếu model có cột ten_phong_ban hoặc ma_phong_ban dạng String thì mới dùng ilike
        if hasattr(ThongTinNguoiLaoDong, 'ten_phong_ban'):
            conditions.append(ThongTinNguoiLaoDong.ten_phong_ban.ilike(f"%{query_str}%"))
        elif hasattr(ThongTinNguoiLaoDong, 'ma_phong_ban'):
            conditions.append(ThongTinNguoiLaoDong.ma_phong_ban.ilike(f"%{query_str}%"))

        base_query = base_query.filter(or_(*conditions))

    total_count = base_query.count()
    records = base_query.offset(offset).limit(per_page).all()

    items = []
    for cb in records:
        ho_ten = getattr(cb, 'ho_ten', '') or cb.ma_nhan_vien
        # Tự động lấy tên phòng ban an toàn dù phong_ban là property hay relationship
        ten_pb = getattr(cb, 'ten_phong_ban', None) or getattr(cb, 'phong_ban', 'N/A')
        if hasattr(ten_pb, 'ten_phong_ban'):  # Nếu phong_ban là một Object Relationship
            ten_pb = ten_pb.ten_phong_ban

        items.append({
            "id": cb.ma_nhan_vien,
            "text": f"{cb.ma_nhan_vien} - {ho_ten} ({ten_pb})"
        })

    has_more = (offset + per_page) < total_count

    return jsonify({
        "items": items,
        "has_more": has_more
    })


# Route cho API Lĩnh vực
@phan_cong_bp.route('/api/search-linh-vuc', methods=['GET'])
@login_required
def search_linh_vuc():
    query = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 20
    offset = (page - 1) * per_page

    base_query = LinhVuc.query
    if query:
        base_query = base_query.filter(LinhVuc.ten_linh_vuc.ilike(f"%{query}%"))

    total_count = base_query.count()
    records = base_query.offset(offset).limit(per_page).all()

    items = [{"id": lv.ten_linh_vuc, "text": lv.ten_linh_vuc} for lv in records]
    has_more = (offset + per_page) < total_count

    return jsonify({"items": items, "has_more": has_more})

# ----------------------------------------------------------------------
# 2. TẠO MỚI PHÂN CÔNG CÔNG TÁC (TRANG RIÊNG BIỆT)
# ----------------------------------------------------------------------
@phan_cong_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create():
    if request.method == 'GET':
        ds_phong_ban = PhongBan.query.all() if hasattr(PhongBan, 'query') else []
        ds_linh_vuc = LinhVuc.query.order_by(LinhVuc.ten_linh_vuc.asc()).all() if hasattr(LinhVuc, 'query') else []
        return render_template(
            'phan_cong_cong_tac/create.html',
            ds_phong_ban=ds_phong_ban,
            ds_linh_vuc=ds_linh_vuc
        )

    try:
        ma_nhan_vien = request.form.get('ma_nhan_vien', '').strip()
        noi_cong_tac = request.form.get('noi_cong_tac', '').strip()

        # Nhận danh sách Lĩnh vực được chọn (Select2 multiple)
        linh_vuc_list = request.form.getlist('linh_vuc') 
        # Nhận nội dung bổ sung gõ tự do
        noi_dung_bo_sung = request.form.get('noi_dung_bo_sung', '').strip()
        
        ngay_bat_dau_str = request.form.get('ngay_bat_dau', '').strip()
        ngay_ket_thuc_str = request.form.get('ngay_ket_thuc', '').strip()
        phuong_tien = request.form.get('phuong_tien', 'Ô tô').strip()
        kinh_phi_du_tru = request.form.get('kinh_phi_du_tru', 0, type=float)
        ghi_chu = request.form.get('ghi_chu', '').strip()

        if not ma_nhan_vien or not noi_cong_tac or (not linh_vuc_list and not noi_dung_bo_sung) or not ngay_bat_dau_str or not ngay_ket_thuc_str:
            flash('Vui lòng nhập đầy đủ cán bộ, nơi công tác, ngày tháng và nội dung công tác!', 'danger')
            return redirect(url_for('phan_cong.create'))

        # Tổng hợp thành chuỗi lưu vào cột noi_dung_cong_tac
        noi_dung_parts = []
        if linh_vuc_list:
            noi_dung_parts.append("Lĩnh vực: " + ", ".join(linh_vuc_list))
        if noi_dung_bo_sung:
            noi_dung_parts.append(f"Chi tiết: {noi_dung_bo_sung}")        
        noi_dung_cong_tac = "\n".join(noi_dung_parts)

        ngay_bat_dau = datetime.strptime(ngay_bat_dau_str, '%Y-%m-%d').date()
        ngay_ket_thuc = datetime.strptime(ngay_ket_thuc_str, '%Y-%m-%d').date()

        if ngay_ket_thuc < ngay_bat_dau:
            flash('Ngày kết thúc không thể trước ngày bắt đầu!', 'warning')
            return redirect(url_for('phan_cong.create'))

        # Lấy trực tiếp thông tin từ ThongTinNguoiLaoDong
        tt_lao_dong = ThongTinNguoiLaoDong.query.filter_by(ma_nhan_vien=ma_nhan_vien).first()
        if tt_lao_dong:
            ho_ten = getattr(tt_lao_dong, 'ho_ten', ma_nhan_vien)
            ten_phong_ban = getattr(tt_lao_dong, 'phong_ban', '')
            ma_phong_ban = getattr(tt_lao_dong, 'ma_phong_ban', '')
        else:
            user_info = User.query.filter_by(ma_nhan_vien=ma_nhan_vien).first()
            ho_ten = getattr(user_info, 'fullname', ma_nhan_vien) if user_info else ma_nhan_vien
            ten_phong_ban = ""
            ma_phong_ban = ""

        trang_thai = 'DA_DUYET' if (getattr(current_user, 'is_admin', False) or getattr(current_user, 'can_approve', False)) else 'CHO_DUYET'

        moi = PhanCongCongTac(
            ma_nhan_vien=ma_nhan_vien,
            ho_ten=ho_ten,
            ma_phong_ban=ma_phong_ban,
            ten_phong_ban=str(ten_phong_ban),
            noi_cong_tac=noi_cong_tac,
            noi_dung_cong_tac=noi_dung_cong_tac,
            ngay_bat_dau=ngay_bat_dau,
            ngay_ket_thuc=ngay_ket_thuc,
            phuong_tien=phuong_tien,
            kinh_phi_du_tru=kinh_phi_du_tru,
            trang_thai=trang_thai,
            nguoi_tao=current_user.ma_nhan_vien,
            ghi_chu=ghi_chu
        )

        db.session.add(moi)
        db.session.commit()
        flash('Đã thêm phân công công tác thành công!', 'success')
        return redirect(url_for('phan_cong.index'))

    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi khi tạo phân công công tác: {str(e)}', 'danger')
        return redirect(url_for('phan_cong.create'))

# ----------------------------------------------------------------------
# 3. CẬP NHẬT PHÂN CÔNG CÔNG TÁC
# ----------------------------------------------------------------------
@phan_cong_bp.route('/update/<int:id>', methods=['POST'])
@login_required
def update(id):
    item = PhanCongCongTac.query.get_or_404(id)

    if not (getattr(current_user, 'is_admin', False) or getattr(current_user, 'can_approve', False) or item.nguoi_tao == current_user.ma_nhan_vien):
        flash('Bạn không có quyền chỉnh sửa bản ghi này!', 'danger')
        return redirect(url_for('phan_cong.index'))

    try:
        item.noi_cong_tac = request.form.get('noi_cong_tac', item.noi_cong_tac).strip()
        item.noi_dung_cong_tac = request.form.get('noi_dung_cong_tac', item.noi_dung_cong_tac).strip()
        
        d_start = request.form.get('ngay_bat_dau')
        d_end = request.form.get('ngay_ket_thuc')

        start_date = datetime.strptime(d_start, '%Y-%m-%d').date() if d_start else item.ngay_bat_dau
        end_date = datetime.strptime(d_end, '%Y-%m-%d').date() if d_end else item.ngay_ket_thuc

        if end_date < start_date:
            flash('Ngày kết thúc không thể trước ngày bắt đầu!', 'warning')
            return redirect(url_for('phan_cong.index'))

        item.ngay_bat_dau = start_date
        item.ngay_ket_thuc = end_date
        item.phuong_tien = request.form.get('phuong_tien', item.phuong_tien)
        item.kinh_phi_du_tru = request.form.get('kinh_phi_du_tru', item.kinh_phi_du_tru, type=float)
        item.ghi_chu = request.form.get('ghi_chu', item.ghi_chu)

        db.session.commit()
        flash('Cập nhật phân công công tác thành công!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi cập nhật: {str(e)}', 'danger')

    return redirect(url_for('phan_cong.index'))

# ----------------------------------------------------------------------
# 4. PHÊ DUYỆT / TỪ CHỐI
# ----------------------------------------------------------------------
@phan_cong_bp.route('/approve/<int:id>', methods=['POST'])
@login_required
def approve(id):
    if not (getattr(current_user, 'is_admin', False) or getattr(current_user, 'can_approve', False)):
        return jsonify({'success': False, 'message': 'Bạn không có quyền duyệt!'}), 403

    item = PhanCongCongTac.query.get_or_404(id)
    action = request.form.get('action') # 'approve', 'reject', hoặc 'complete'
    ly_do = request.form.get('ly_do_tu_choi', '')

    try:
        if action == 'approve':
            item.trang_thai = 'DA_DUYET'
            item.ly_do_tu_choi = None
        elif action == 'reject':
            item.trang_thai = 'TU_CHOI'
            item.ly_do_tu_choi = ly_do
        elif action == 'complete':
            item.trang_thai = 'HOAN_THANH'

        item.nguoi_duyet = current_user.ma_nhan_vien
        item.ngay_duyet = datetime.now()

        db.session.commit()
        return jsonify({'success': True, 'message': 'Thao tác phê duyệt thành công!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ----------------------------------------------------------------------
# 5. XÓA PHÂN CÔNG CÔNG TÁC
# ----------------------------------------------------------------------
@phan_cong_bp.route('/delete/<int:id>', methods=['POST'])
@login_required
def delete(id):
    item = PhanCongCongTac.query.get_or_404(id)

    if not (getattr(current_user, 'is_admin', False) or item.nguoi_tao == current_user.ma_nhan_vien):
        flash('Bạn không có quyền xóa bản ghi này!', 'danger')
        return redirect(url_for('phan_cong.index'))

    try:
        db.session.delete(item)
        db.session.commit()
        flash('Đã xóa lịch phân công công tác!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi khi xóa: {str(e)}', 'danger')

    return redirect(url_for('phan_cong.index'))

# ----------------------------------------------------------------------
# 6. XUẤT EXCEL DANH SÁCH
# ----------------------------------------------------------------------
@phan_cong_bp.route('/export', methods=['GET'])
@login_required
def export_excel():
    keyword = request.args.get('keyword', '').strip()
    trang_thai = request.args.get('trang_thai', '').strip()
    tu_ngay = request.args.get('tu_ngay', '').strip()
    den_ngay = request.args.get('den_ngay', '').strip()

    query = PhanCongCongTac.query

    # Phân quyền tương tự trang index
    if not (getattr(current_user, 'is_admin', False) or getattr(current_user, 'can_approve', False) or getattr(current_user, 'is_controller', False)):
        query = query.filter(
            or_(
                PhanCongCongTac.ma_nhan_vien == current_user.ma_nhan_vien,
                PhanCongCongTac.nguoi_tao == current_user.ma_nhan_vien
            )
        )

    if keyword:
        query = query.filter(
            or_(
                PhanCongCongTac.ho_ten.ilike(f'%{keyword}%'),
                PhanCongCongTac.ma_nhan_vien.ilike(f'%{keyword}%'),
                PhanCongCongTac.noi_cong_tac.ilike(f'%{keyword}%'),
                PhanCongCongTac.noi_dung_cong_tac.ilike(f'%{keyword}%')
            )
        )

    if trang_thai:
        query = query.filter(PhanCongCongTac.trang_thai == trang_thai)

    if tu_ngay:
        try:
            d_tu = datetime.strptime(tu_ngay, '%Y-%m-%d').date()
            query = query.filter(PhanCongCongTac.ngay_bat_dau >= d_tu)
        except ValueError:
            pass

    if den_ngay:
        try:
            d_den = datetime.strptime(den_ngay, '%Y-%m-%d').date()
            query = query.filter(PhanCongCongTac.ngay_ket_thuc <= d_den)
        except ValueError:
            pass

    items = query.order_by(PhanCongCongTac.ngay_bat_dau.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PhanCongCongTac"

    font_header = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    headers = [
        "STT", "Mã NV", "Họ và Tên", "Phòng Ban", "Nơi Công Tác", 
        "Nội Dung Công Tác", "Ngày Bắt Đầu", "Ngày Kết Thúc", 
        "Phương Tiện", "Kinh Phí Dự Trù", "Trạng Thái", "Ghi Chú"
    ]

    ws.append(headers)
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    for idx, item in enumerate(items, 1):
        row = [
            idx,
            item.ma_nhan_vien,
            item.ho_ten,
            item.ten_phong_ban or "",
            item.noi_cong_tac,
            item.noi_dung_cong_tac,
            item.ngay_bat_dau.strftime('%d/%m/%Y') if item.ngay_bat_dau else "",
            item.ngay_ket_thuc.strftime('%d/%m/%Y') if item.ngay_ket_thuc else "",
            item.phuong_tien or "",
            item.kinh_phi_du_tru or 0,
            item.trang_thai,
            item.ghi_chu or ""
        ]
        ws.append(row)
        
        current_row = idx + 1
        for col_num in range(1, len(headers) + 1):
            c = ws.cell(row=current_row, column=col_num)
            c.font = Font(name="Times New Roman", size=11)
            c.border = thin_border
            if col_num in [1, 2, 7, 8, 11]:
                c.alignment = align_center

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Danh_sach_phan_cong_cong_tac_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)

# ----------------------------------------------------------------------
# 7. API XEM CHI TIẾT
# ----------------------------------------------------------------------
@phan_cong_bp.route('/detail/<int:id>', methods=['GET'])
@login_required
def detail(id):
    item = PhanCongCongTac.query.get_or_404(id)
    return jsonify({
        'success': True,
        'data': {
            'id': item.id,
            'ma_nhan_vien': item.ma_nhan_vien,
            'ho_ten': item.ho_ten,
            'ten_phong_ban': item.ten_phong_ban or '',
            'noi_cong_tac': item.noi_cong_tac,
            'noi_dung_cong_tac': item.noi_dung_cong_tac,
            'ngay_bat_dau': item.ngay_bat_dau.strftime('%d/%m/%Y') if item.ngay_bat_dau else '',
            'ngay_ket_thuc': item.ngay_ket_thuc.strftime('%d/%m/%Y') if item.ngay_ket_thuc else '',
            'phuong_tien': item.phuong_tien or '',
            'kinh_phi_du_tru': item.kinh_phi_du_tru or 0,
            'trang_thai': item.trang_thai,
            'ly_do_tu_choi': item.ly_do_tu_choi or '',
            'nguoi_tao': item.nguoi_tao or '',
            'nguoi_duyet': item.nguoi_duyet or '',
            'ghi_chu': item.ghi_chu or ''
        }
    })

# ----------------------------------------------------------------------
# 8. BÁO CÁO & THỐNG KÊ
# ----------------------------------------------------------------------
@phan_cong_bp.route('/thong-ke', methods=['GET'])
@login_required
def thong_ke():
    tong_so = db.session.query(func.count(PhanCongCongTac.id)).scalar() or 0
    cho_duyet = db.session.query(func.count(PhanCongCongTac.id)).filter(PhanCongCongTac.trang_thai == 'CHO_DUYET').scalar() or 0
    da_duyet = db.session.query(func.count(PhanCongCongTac.id)).filter(PhanCongCongTac.trang_thai == 'DA_DUYET').scalar() or 0
    hoan_thanh = db.session.query(func.count(PhanCongCongTac.id)).filter(PhanCongCongTac.trang_thai == 'HOAN_THANH').scalar() or 0
    tong_kinh_phi = db.session.query(func.sum(PhanCongCongTac.kinh_phi_du_tru)).scalar() or 0

    by_dept = db.session.query(
        PhanCongCongTac.ten_phong_ban,
        func.count(PhanCongCongTac.id).label('so_luong'),
        func.sum(PhanCongCongTac.kinh_phi_du_tru).label('tong_kinh_phi')
    ).group_by(PhanCongCongTac.ten_phong_ban).all()

    return render_template(
        'phan_cong_cong_tac/thong_ke.html',
        tong_so=tong_so,
        cho_duyet=cho_duyet,
        da_duyet=da_duyet,
        hoan_thanh=hoan_thanh,
        tong_kinh_phi=tong_kinh_phi,
        by_dept=by_dept
    )
